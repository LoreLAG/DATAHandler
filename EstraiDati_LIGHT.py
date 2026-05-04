import pdfplumber, re, csv, threading
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
from pathlib import Path
import ctypes

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# ── COSTANTI GRAFICHE ────────────────────────────────────────────────────────
CHECKED = "☑"
UNCHECKED = "☐"

# ── REGEX MOTORE ─────────────────────────────────────────────────────────────
RE_CALYPSO = re.compile(
    r"^(.+?)[ \t]+(-?\d+[,\.]\d+)[ \t]*mm[ \t]+(-?\d+[,\.]\d+)[ \t]+(-?\d+[,\.]\d+)[ \t]+(-?\d+[,\.]\d+)[ \t]+(-?\d+[,\.]\d+)(?:[ \t]+(-?\d+[,\.]\d+))?$")
RE_SECTION = re.compile(r"^(V\d+_F\d+)$")
RE_ANGLE = re.compile(r"^(.+?)[ \t]+(-?\d+[,\.]\d+)°[ \t]+(-?\d+[,\.]\d+)[ \t]+(-?\d+[,\.]\d+)$")
RE_PARTIAL = re.compile(r"^(.+?)[ \t]+(-?\d+[,\.]\d+)[ \t]*mm[ \t]*$")


def parse_number(s):
    if not s: return None
    try:
        return float(s.strip().replace(" mm", "").replace(",", "."))
    except ValueError:
        return None


def pulisci_testo(testo):
    return " ".join(str(testo).split()) if testo else ""


# ── LOGICA ESTRAZIONE ────────────────────────────────────────────────────────
def extract_single_pdf(pdf_path, has_sections=False):
    rows = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            current_section = ""
            seen = set()
            for page in pdf.pages:
                text = page.extract_text()
                if not text: continue
                for line in text.splitlines():
                    line = line.strip()
                    if has_sections and RE_SECTION.match(line):
                        current_section = RE_SECTION.match(line).group(1)
                        continue

                    row = None
                    m = RE_CALYPSO.match(line)
                    if m:
                        g = m.groups()
                        tp, tm, dev = parse_number(g[3]), parse_number(g[4]), parse_number(g[5])
                        stato = "OK" if (
                                tm is not None and tp is not None and dev is not None and tm <= dev <= tp) else "NON OK"
                        row = {"section": current_section, "name": pulisci_testo(g[0]), "measured": parse_number(g[1]),
                               "nominal": parse_number(g[2]), "tp": tp, "tm": tm, "dev": dev, "stato": stato}
                    else:
                        ma = RE_ANGLE.match(line)
                        if ma:
                            g = ma.groups()
                            row = {"section": current_section, "name": pulisci_testo(g[0]),
                                   "measured": parse_number(g[1]), "nominal": parse_number(g[2]), "tp": None,
                                   "tm": None, "dev": parse_number(g[3]), "stato": None}
                        else:
                            mp = RE_PARTIAL.match(line)
                            if mp:
                                g = mp.groups()
                                row = {"section": current_section, "name": pulisci_testo(g[0]),
                                       "measured": parse_number(g[1]), "nominal": None, "tp": None, "tm": None,
                                       "dev": None, "stato": None}

                    if row and (current_section, row["name"]) not in seen:
                        seen.add((current_section, row["name"]))
                        rows.append(row)
    except Exception as e:
        print(f"Errore file {pdf_path}: {e}")
    return rows


# ── INTERFACCIA ──────────────────────────────────────────────────────────────
class CalypsoLightApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Calypso Data Extractor - Light Edition")
        self.root.geometry("1300x750")
        self.root.minsize(900, 600)

        self.output_dir = tk.StringVar(value=str(Path.home() / "Desktop" / "Estrazioni_Calypso"))
        self.use_sections = tk.BooleanVar(value=True)

        self.setup_ui()

    def setup_ui(self):
        style = ttk.Style()
        style.theme_use("clam")

        style.configure("Treeview", rowheight=28, font=("Arial", 10))
        style.configure("Treeview.Heading", font=("Arial", 10, "bold"), background="#d5d8dc")

        # ==========================================
        # 1. TOP BAR (Stile ConsultaDB)
        # ==========================================
        frame_top = tk.Frame(self.root, bg="#1c2833", pady=15, padx=10)
        frame_top.pack(fill="x", side="top")

        tk.Label(frame_top, text="📁 Gestione Sorgenti (PDF)", bg="#1c2833", fg="white",
                 font=("Arial", 14, "bold")).pack(side="left", padx=10)

        self.btn_clear = tk.Button(frame_top, text="🗑 Svuota Albero", command=self.clear_list,
                                   bg="#e74c3c", fg="white", font=("Arial", 10, "bold"), bd=0, padx=10, pady=5)
        self.btn_clear.pack(side="right", padx=5)

        self.btn_add_files = tk.Button(frame_top, text="📄 Aggiungi PDF Singoli", command=self.add_files,
                                       bg="#3498db", fg="white", font=("Arial", 10, "bold"), bd=0, padx=10, pady=5)
        self.btn_add_files.pack(side="right", padx=5)

        self.btn_add_folder = tk.Button(frame_top, text="📁 Aggiungi Cartella (Ricorsiva)", command=self.add_folder,
                                        bg="#27ae60", fg="white", font=("Arial", 10, "bold"), bd=0, padx=10, pady=5)
        self.btn_add_folder.pack(side="right", padx=5)

        # ==========================================
        # 2. BODY: TREEVIEW A CASCATA
        # ==========================================
        frame_body = tk.Frame(self.root, padx=20, pady=15)
        frame_body.pack(fill="both", expand=True)

        # Torniamo all'impostazione nativa con "path" mostrato direttamente
        self.tree = ttk.Treeview(frame_body, columns=("count", "path"), show="tree headings")

        self.tree.heading("#0", text="Esporta | Nome Cartella / File")
        self.tree.heading("count", text="PDF")
        self.tree.heading("path", text="Percorso Completo")

        # Larghezze ottimizzate, percorso riallineato a sinistra ("w")
        self.tree.column("#0", width=280, anchor="w")
        self.tree.column("count", width=40, anchor="center")
        self.tree.column("path", width=550, anchor="w")

        scrollbar = ttk.Scrollbar(frame_body, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        self.tree.pack(side="left", fill="both", expand=True)

        self.tree.bind("<ButtonRelease-1>", self.toggle_check)

        # ==========================================
        # 3. BOTTOM BAR E LOG
        # ==========================================
        frame_bot = tk.Frame(self.root, padx=20, pady=10, bg="#ecf0f1")
        frame_bot.pack(fill="x", side="bottom")

        # Sotto-sezione destinazione
        frame_dest = tk.Frame(frame_bot, bg="#ecf0f1")
        frame_dest.pack(fill="x", pady=(0, 10))

        tk.Label(frame_dest, text="Cartella di Destinazione Output:", font=("Arial", 10, "bold"), bg="#ecf0f1").pack(
            side="left")
        tk.Entry(frame_dest, textvariable=self.output_dir, width=70, font=("Arial", 10)).pack(side="left", padx=10)
        tk.Button(frame_dest, text="Sfoglia...", command=self.browse_out, font=("Arial", 9)).pack(side="left")

        # Sotto-sezione Azioni
        self.btn_run = tk.Button(frame_bot, text="🚀 AVVIA AGGREGAZIONE DATI", bg="#27ae60", fg="white",
                                 font=("Arial", 14, "bold"), height=2, bd=0, command=self.start_process)
        self.btn_run.pack(fill="x", pady=(5, 5))

        self.progress = ttk.Progressbar(frame_bot, orient="horizontal", mode="determinate")
        self.progress.pack(fill="x", pady=(5, 10))

        # Log
        self.log_box = scrolledtext.ScrolledText(frame_bot, height=6, font=("Consolas", 9), bg="#1e1e1e", fg="#00ff00")
        self.log_box.pack(fill="both", expand=True)

    # ── HANDLERS COSTRUZIONE ALBERO ──────────────────────────────────────────
    def build_tree_node(self, parent_id, path_obj, depth=0):
        # Limite profondità per non appesantire l'albero visivo
        if depth > 2:
            return 0

        pdfs = list(path_obj.glob("*.pdf"))
        pdf_count = len(pdfs)

        node_name = path_obj.name

        # Inseriamo il nodo: nella prima colonna il nome, nell'ultima il percorso completo stringa
        node_id = self.tree.insert(parent_id, "end", text=f"{CHECKED}  {node_name}", values=(pdf_count, str(path_obj)))

        total_pdfs_in_branch = pdf_count

        try:
            for sub in path_obj.iterdir():
                if sub.is_dir():
                    sub_count = self.build_tree_node(node_id, sub, depth + 1)
                    total_pdfs_in_branch += sub_count
        except PermissionError:
            pass

        # Pruning dei rami vuoti
        if total_pdfs_in_branch == 0:
            self.tree.delete(node_id)
            return 0

        if depth == 0:
            self.tree.item(node_id, open=True)

        return total_pdfs_in_branch

    def add_folder(self):
        d = filedialog.askdirectory(parent=self.root, title="Seleziona cartella")
        if d:
            path_obj = Path(d)
            self.log(f"Scansione in corso per: {path_obj.name}...")
            tot = self.build_tree_node("", path_obj, depth=0)
            if tot == 0:
                messagebox.showwarning("Vuota", "Nessun PDF trovato nella cartella o nelle sue sottocartelle valide.",
                                       parent=self.root)
            else:
                self.log(f"Aggiunta struttura ad albero ({tot} PDF totali nel ramo).")

    def add_files(self):
        fs = filedialog.askopenfilenames(parent=self.root, filetypes=[("PDF Files", "*.pdf")])
        if fs:
            for f in fs:
                path_obj = Path(f)
                self.tree.insert("", "end", text=f"{CHECKED}  {path_obj.name}", values=(1, str(path_obj)))
            self.log(f"Aggiunti {len(fs)} file PDF singoli.")

    # ── LOGICA SPUNTE A CASCATA ──────────────────────────────────────────────
    def set_node_state(self, item_id, state_char):
        current_text = self.tree.item(item_id, "text")
        base_name = current_text[1:].strip()
        self.tree.item(item_id, text=f"{state_char}  {base_name}")

        for child in self.tree.get_children(item_id):
            self.set_node_state(child, state_char)

    def toggle_check(self, event):
        item_id = self.tree.identify_row(event.y)
        if not item_id: return

        elemento = self.tree.identify_element(event.x, event.y)
        if "indicator" in elemento:
            return

        column = self.tree.identify_column(event.x)
        if column == "#0":
            bbox = self.tree.bbox(item_id, "#0")
            if not bbox: return

            x_inizio_testo = bbox[0]
            click_relativo = event.x - x_inizio_testo

            if click_relativo > 25 or click_relativo < 0:
                return

            current_text = self.tree.item(item_id, "text")
            if current_text.startswith(CHECKED):
                self.set_node_state(item_id, UNCHECKED)
            elif current_text.startswith(UNCHECKED):
                self.set_node_state(item_id, CHECKED)

    def clear_list(self):
        for i in self.tree.get_children(): self.tree.delete(i)
        self.log("Lista sorgenti svuotata.")

    def browse_out(self):
        d = filedialog.askdirectory(parent=self.root, title="Seleziona la cartella di destinazione")
        if d: self.output_dir.set(d)

    def log(self, msg):
        self.log_box.config(state='normal')
        self.log_box.insert(tk.END, f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n")
        self.log_box.see(tk.END)
        self.log_box.config(state='disabled')

    # ── MOTORE DI ELABORAZIONE ───────────────────────────────────────────────
    def get_checked_paths(self, node):
        paths = []
        if self.tree.item(node, "text").startswith(CHECKED):
            # values[1] contiene str(path_obj) ovvero il percorso reale
            paths.append(self.tree.item(node, "values")[1])

        for child in self.tree.get_children(node):
            paths.extend(self.get_checked_paths(child))
        return paths

    def start_process(self):
        selected_paths = []
        for root_node in self.tree.get_children():
            selected_paths.extend(self.get_checked_paths(root_node))

        if not selected_paths:
            messagebox.showwarning("Attenzione", "Nessuna cartella spuntata da elaborare.", parent=self.root)
            return

        self.btn_run.config(state="disabled", text="ELABORAZIONE IN CORSO...")
        self.btn_add_folder.config(state="disabled")
        self.btn_add_files.config(state="disabled")
        self.btn_clear.config(state="disabled")
        threading.Thread(target=self.process_thread, args=(selected_paths,), daemon=True).start()

    def process_thread(self, selected_paths):
        try:
            out_path = Path(self.output_dir.get())
            out_path.mkdir(parents=True, exist_ok=True)

            all_data = []
            pdf_tasks = []

            for s in selected_paths:
                p = Path(s)
                if p.is_dir():
                    pdf_tasks.extend(list(p.glob("*.pdf")))
                elif p.is_file():
                    pdf_tasks.append(p)

            pdf_tasks = list(set(pdf_tasks))
            total = len(pdf_tasks)

            if total == 0:
                self.log("Nessun PDF valido trovato nelle selezioni.")
                return

            self.root.after(0, lambda: self.progress.configure(maximum=total, value=0))

            for idx, pdf_path in enumerate(pdf_tasks):
                self.log(f"Analisi {pdf_path.name}...")
                results = extract_single_pdf(pdf_path, self.use_sections.get())
                for r in results:
                    r["codice_pezzo"] = pdf_path.stem
                    all_data.append(r)

                self.root.after(0, lambda v=idx + 1: self.progress.configure(value=v))

            if all_data:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M")
                headers = ["codice_pezzo", "section", "name", "measured", "nominal", "tp", "tm", "dev", "stato"]

                csv_file = out_path / f"Aggregato_{timestamp}.csv"
                with open(csv_file, "w", newline="", encoding="utf-8-sig") as f:
                    w = csv.DictWriter(f, fieldnames=headers, delimiter=";")
                    w.writeheader()
                    w.writerows(all_data)

                if EXCEL_AVAILABLE:
                    xlsx_file = out_path / f"Aggregato_{timestamp}.xlsx"
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.append(headers)
                    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    for r in all_data:
                        ws.append([r.get(h) for h in headers])
                        if r.get("stato") == "NON OK":
                            for cell in ws[ws.max_row]: cell.fill = red_fill
                    wb.save(xlsx_file)
                    self.log(f"Salvataggio Excel completato: {xlsx_file.name}")

                self.log(f"✅ FINITO! Estratte {len(all_data)} misure da {total} file.")
                self.root.after(0, lambda: messagebox.showinfo("Successo",
                                                               f"Elaborazione completata.\nFile salvati in: {out_path}",
                                                               parent=self.root))
            else:
                self.log("⚠️ Nessun dato estratto dai file selezionati.")

        except Exception as e:
            self.log(f"❌ ERRORE CRITICO: {e}")
        finally:
            self.root.after(0, lambda: [
                self.btn_run.config(state="normal", text="AVVIA ESTRAZIONE DATI"),
                self.btn_add_folder.config(state="normal"),
                self.btn_add_files.config(state="normal"),
                self.btn_clear.config(state="normal"),
                self.progress.configure(value=0)
            ])


if __name__ == "__main__":
    try:
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("Calypso.Extractor.Light")
    except Exception:
        pass

    root = tk.Tk()
    app = CalypsoLightApp(root)
    root.mainloop()